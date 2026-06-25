#!/usr/bin/env python3
"""
ETF 份额数据采集脚本
支持本地 Windows/Mac 运行 + GitHub Actions 自动运行
"""

import os, sys, time, re
import requests
import pandas as pd
from io import BytesIO
from datetime import datetime, date, timedelta
from chinese_calendar import is_workday
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# ================= 配置区 =================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if sys.platform == "win32":
    DEFAULT_EXCEL = r"D:\00python\ETF份额统计.xlsx"
    if not os.path.exists(DEFAULT_EXCEL):
        DEFAULT_EXCEL = os.path.join(SCRIPT_DIR, "assets", "posts", "ETF份额统计.xlsx")
else:
    DEFAULT_EXCEL = os.path.join(SCRIPT_DIR, "assets", "posts", "ETF份额统计.xlsx")

EXCEL_FILE_PATH = DEFAULT_EXCEL

SSE_LIST = [
    "510300", "510310", "510050", "510500", "512100", "512400", "516650",
    "517520", "512480", "588200", "588000", "515070", "512660", "562500",
    "512170", "512010", "515790", "516970", "512200", "516510", "512880",
    "512800", "513090", "513180", "516150", "512690", "515220", "561360",
    "510170", "512890", "560080", "511130", "511260"
]

SZSE_LIST = [
    '159915', '159919', '159326', '159852', '159870',
    '159516', '159819', '159206', '159992', '159755',
    '159745', '159611', '159851', '159920', '159567', '159928', '159998', '159985'
]

SEARCH_ROW      = 13
DATA_START_ROW  = 18
DATE_COL        = 3
MAX_HISTORY_ROW = 163

ROW_QUARTERLY = 7
ROW_MONTHLY   = 8
ROW_WEEKLY    = 9
ROW_DAILY     = 10

ROW_NAV_DAILY   = 15
ROW_NAV_MONTHLY = 14
ROW_CASH_FLOW   = 16

TEST_MODE = False
# ==========================================


def get_beijing_date():
    return (datetime.utcnow() + timedelta(hours=8)).date()


def check_trading_day():
    if TEST_MODE:
        print("⚠️  [测试模式] 已跳过交易日检查，强制运行")
        return True
    bj_today  = get_beijing_date()
    yesterday = bj_today - timedelta(days=1)
    if not is_workday(yesterday):
        bj_now = datetime.utcnow() + timedelta(hours=8)
        print(f"{'='*55}")
        print(f"  ETF 份额采集  {bj_now.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*55}")
        print(f"\n⏭️  昨天 {yesterday} 非A股交易日，今日无需运行，退出。")
        print(f"{'='*55}")
        return False
    return True


def get_next_trading_day(d):
    next_day = d + timedelta(days=1)
    while not is_workday(next_day):
        next_day += timedelta(days=1)
    return next_day


def is_last_trading_day_of_week(d):
    return get_next_trading_day(d).isocalendar()[1] != d.isocalendar()[1]


def is_last_trading_day_of_month(d):
    return get_next_trading_day(d).month != d.month


def is_last_trading_day_of_quarter(d):
    if d.month not in [3, 6, 9, 12]:
        return False
    return is_last_trading_day_of_month(d)


def get_market_value_target_row(yesterday):
    if is_last_trading_day_of_quarter(yesterday):
        print(f"   📊 昨天 {yesterday} 是季度最后交易日 → 市值填入第{ROW_QUARTERLY}行")
        return ROW_QUARTERLY
    elif is_last_trading_day_of_month(yesterday):
        print(f"   📊 昨天 {yesterday} 是月份最后交易日 → 市值填入第{ROW_MONTHLY}行")
        return ROW_MONTHLY
    elif is_last_trading_day_of_week(yesterday):
        print(f"   📊 昨天 {yesterday} 是本周最后交易日 → 市值填入第{ROW_WEEKLY}行")
        return ROW_WEEKLY
    else:
        print(f"   📊 昨天 {yesterday} 是普通交易日 → 市值填入第{ROW_DAILY}行")
        return ROW_DAILY


def shift_formula_rows(formula, delta):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    def replace_row(match):
        return f"{match.group(1)}{int(match.group(2)) + delta}"
    return re.sub(r'([A-Z]+)(\d+)', replace_row, formula)


def prepare_excel_new_row():
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"❌ 找不到 Excel: {EXCEL_FILE_PATH}")
        return None
    try:
        with open(EXCEL_FILE_PATH, 'a'):
            pass
    except IOError:
        print(f"❌ 文件正在被占用！请先关闭 Excel 后再运行。")
        return None
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
        print(f"📋 工作表: {ws.title}")

        code_to_col = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(SEARCH_ROW, col).value
            if val is not None:
                try:
                    code = str(int(float(str(val))))
                    if len(code) == 6:
                        code_to_col[code] = col
                except:
                    pass
        print(f"   🔍 找到 {len(code_to_col)} 个ETF代码映射")

        row18_snapshot = {
            col: ws.cell(DATA_START_ROW, col).value
            for col in range(1, ws.max_column + 1)
        }

        for row in range(MAX_HISTORY_ROW, DATA_START_ROW - 1, -1):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row, col).value
                if isinstance(val, str) and val.startswith('='):
                    ws.cell(row + 1, col).value = shift_formula_rows(val, +1)
                else:
                    ws.cell(row + 1, col).value = val

        for col in range(1, ws.max_column + 1):
            ws.cell(DATA_START_ROW, col).value = None
        for col, val in row18_snapshot.items():
            if isinstance(val, str) and val.startswith('='):
                ws.cell(DATA_START_ROW, col).value = val

        wb.save(EXCEL_FILE_PATH)
        print(f"   ✅ 第18~{MAX_HISTORY_ROW}行整块下移完成")
        print(f"   ✅ 第18行已清空并还原公式，等待数据填入")
        return code_to_col

    except PermissionError:
        print(f"❌ 文件写入被拒绝，请确认 Excel 已完全关闭后重试")
        return None
    except Exception as e:
        print(f"❌ Excel 准备失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def fetch_szse_data(code_list):
    """
    深交所 ETF 数据（最多重试3次）
    触发重试条件：①响应内容 <1000 bytes  ②份额或涨跌幅有缺漏
    返回 (share_map, market_map, nav_map, change_map)
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Referer": "https://fund.szse.cn/",
    }

    results_map = {}
    market_map  = {}
    nav_map     = {}
    change_map  = {}

    for attempt in range(1, 4):
        try:
            url = (
                "https://fund.szse.cn/api/report/ShowReport"
                "?SHOWTYPE=xlsx&CATALOGID=1000_lf&TABKEY=tab1"
                f"&random={time.time()}"
            )
            print(f"\n--- 正在同步深交所数据 (尝试 {attempt}/3) ---")
            response = requests.get(url, headers=headers, timeout=30)

            # ① 响应内容过小 → 重试
            if len(response.content) < 1000:
                print(f"   ⚠️ 第{attempt}次返回为空({len(response.content)} bytes)，重试中...")
                time.sleep(3 * attempt)
                continue

            df = pd.read_excel(BytesIO(response.content))
            df['当前规模(份)'] = (
                df['当前规模(份)'].astype(str).str.replace(',', '').str.strip()
            )
            df['当前规模(份)'] = pd.to_numeric(df['当前规模(份)'], errors='coerce')

            # 自动识别涨跌幅列
            change_col = next(
                (c for c in df.columns if any(k in str(c) for k in ['涨跌幅', '涨跌', '升跌'])),
                None
            )
            if change_col:
                print(f"   📊 深交所涨跌幅列: 「{change_col}」")
            else:
                print(f"   ⚠️ 深交所主表未找到涨跌幅列，列名: {list(df.columns)}")

            # 逐只解析
            tmp_share  = {}
            tmp_market = {}
            tmp_nav    = {}
            tmp_change = {}

            for code in code_list:
                mask       = df['基金代码'].astype(str).str.strip() == str(code)
                target_row = df[mask]
                if not target_row.empty:
                    shares = target_row.iloc[0]['当前规模(份)']
                    nav    = target_row.iloc[0]['净值']
                    if pd.notna(shares) and pd.notna(nav) and shares > 0:
                        market_val = round(shares * nav / 100_000_000, 6)
                        tmp_share[str(code)]  = shares / 10_000
                        tmp_market[str(code)] = market_val
                        tmp_nav[str(code)]    = float(nav)
                        if change_col:
                            raw_chg = target_row.iloc[0][change_col]
                            if pd.notna(raw_chg):
                                try:
                                    tmp_change[str(code)] = round(float(raw_chg) / 100, 6)
                                except:
                                    pass
                        chg_str = f"{tmp_change[str(code)]*100:.2f}%" if str(code) in tmp_change else "涨跌幅未获取"
                        print(f"   ✅ 深交所 {code}: {shares/10000:.2f} 万份"
                              f"  净值={nav}  涨跌幅={chg_str}  市值={market_val:.4f}亿")
                    else:
                        print(f"   ⚠️ 深交所 {code}: 份额或净值为空")
                else:
                    print(f"   ⚠️ 深交所未找到: {code}")

            # ② 涨跌幅有缺漏 → 重试（份额已有则不重试份额，只重试整体以拿齐涨跌幅）
            missing_change = [c for c in code_list if str(c) not in tmp_change]
            if missing_change and attempt < 3:
                print(f"   ⚠️ 第{attempt}次：{len(missing_change)} 只涨跌幅未获取"
                      f"（{missing_change[:5]}{'...' if len(missing_change) > 5 else ''}），重试中...")
                time.sleep(3 * attempt)
                continue  # 重新拉取整张表

            # 本次成功，写入最终结果
            results_map = tmp_share
            market_map  = tmp_market
            nav_map     = tmp_nav
            change_map  = tmp_change
            break

        except Exception as e:
            if attempt < 3:
                print(f"   ⚠️ 深交所请求失败(第{attempt}次): {e}")
                time.sleep(3 * attempt)
            else:
                print(f"❌ 深交所同步失败(已重试3次): {e}")

    return results_map, market_map, nav_map, change_map


def extract_sse_data_dom(driver, code, max_retries=2):
    """从上交所页面 DOM 提取份额、市值、净值、涨跌幅"""
    for attempt in range(1, max_retries + 1):
        try:
            wait = WebDriverWait(driver, 15)
            wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div.js_fundSize table.table tbody tr")
            ))
            rows = driver.find_elements(
                By.CSS_SELECTOR, "div.js_fundSize table.table tbody tr"
            )
            if rows:
                tds = rows[0].find_elements(By.TAG_NAME, "td")
                if len(tds) >= 4:
                    raw_date  = tds[0].text.strip()
                    raw_share = tds[-1].text.replace(',', '').strip()
                    share_val = float(raw_share)

                    market_val = None
                    nav_val    = None
                    change_val = None

                    # 行情指标区域：第1个td=当日净值（现价），第2个td=涨跌幅
                    try:
                        wait.until(EC.presence_of_element_located(
                            (By.CSS_SELECTOR, "div.js_marketIndex td.colData_val")
                        ))
                        market_index_tds = driver.find_elements(
                            By.CSS_SELECTOR, "div.js_marketIndex td.colData_val"
                        )
                        if market_index_tds:
                            raw_nav = market_index_tds[0].text.replace(',', '').strip()
                            if raw_nav:
                                nav_val = float(raw_nav)
                        # 第2个 td = 涨跌幅（形如 "2.15%"）
                        if len(market_index_tds) >= 2:
                            raw_chg = (market_index_tds[1].text
                                       .replace('%', '').replace(',', '').strip())
                            if raw_chg:
                                change_val = round(float(raw_chg) / 100, 6)
                    except:
                        pass

                    # 成交概览区域：第1个td=基金规模（市值）
                    try:
                        wait.until(EC.presence_of_element_located(
                            (By.CSS_SELECTOR, "div.js_transactionOverview td.colData_val")
                        ))
                        market_tds = driver.find_elements(
                            By.CSS_SELECTOR, "div.js_transactionOverview td.colData_val"
                        )
                        if market_tds:
                            raw_market = market_tds[0].text.replace(',', '').strip()
                            market_val = round(float(raw_market) / 10000, 6)
                    except:
                        pass

                    if share_val > 0:
                        parts    = raw_date.split("-")
                        date_fmt = f"{parts[0]}//{int(parts[1])}/{int(parts[2])}"
                        return share_val, market_val, date_fmt, nav_val, change_val

        except Exception as e:
            if attempt < max_retries:
                time.sleep(2)
            else:
                print(f"   ⚠️ DOM读取失败 {code}: {e}")

    return None, None, None, None, None


def fetch_sse_data(code_list):
    """
    上交所 ETF 数据（Selenium 抓取 DOM）
    返回 (share_map, market_map, detected_date, nav_map, change_map)
    """
    share_map     = {}
    market_map    = {}
    nav_map       = {}
    change_map    = {}
    failed_list   = []
    detected_date = None

    options = Options()
    options.add_argument('--headless')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--log-level=3')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=options
    )

    print(f"\n🚀 开始上交所处理 (共 {len(code_list)} 只)...")
    try:
        for idx, code in enumerate(code_list, start=1):
            try:
                driver.get(
                    f"https://www.sse.com.cn/assortment/fund/list/etfinfo/basic/"
                    f"index.shtml?FUNDID={code}"
                )
                driver.execute_script("window.scrollTo(0, 1800);")
                share, market, date_str, nav, change = extract_sse_data_dom(driver, code)
                if share:
                    share_map[str(code)] = share
                    if market is not None:
                        market_map[str(code)] = market
                    if nav is not None:
                        nav_map[str(code)] = nav
                    if change is not None:
                        change_map[str(code)] = change
                    if detected_date is None and date_str:
                        detected_date = date_str
                    mkt_str = f"{market:.4f}亿" if market else "市值未获取"
                    nav_str = f"{nav}"           if nav    else "净值未获取"
                    chg_str = f"{change*100:.2f}%" if change is not None else "涨跌幅未获取"
                    print(f"   [{idx}/{len(code_list)}] ✅ 沪市 {code}:"
                          f" 份额={share:.2f}万份  净值={nav_str}"
                          f"  涨跌幅={chg_str}  市值={mkt_str}  日期={date_str}")
                else:
                    failed_list.append(code)
                    print(f"   [{idx}/{len(code_list)}] ⚠️ 沪市 {code} 失败")
            except Exception as e:
                failed_list.append(code)
                print(f"   [{idx}/{len(code_list)}] ❌ 沪市 {code} 异常: {e}")
    finally:
        driver.quit()

    if failed_list:
        print(f"\n⚠️ 未获取到数据的沪市ETF: {failed_list}")
    return share_map, market_map, detected_date, nav_map, change_map


def fill_excel_data(code_to_col, share_dict, market_dict, data_date,
                    market_target_row, nav_dict, is_month_end, change_dict):
    """填入数据：份额 + 市值 + 净值 + 资金流动 + 涨跌幅"""
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # ── 日期 ──
        if data_date:
            ws.cell(DATA_START_ROW, DATE_COL).value = data_date
            print(f"\n   📅 日期（来自上交所）: {data_date}")
        else:
            bj_today = get_beijing_date()
            date_fmt = f"{bj_today.year}//{bj_today.month}/{bj_today.day}"
            ws.cell(DATA_START_ROW, DATE_COL).value = date_fmt
            print(f"\n   📅 日期（系统备用）: {date_fmt}")

        # ── 份额 第18行（share_col = market_col + 2）──
        filled_share = 0
        for code, share_val in share_dict.items():
            if code in code_to_col:
                share_col = code_to_col[code] + 2
                ws.cell(DATA_START_ROW, share_col).value = float(share_val)
                filled_share += 1
        print(f"   📈 份额填入完成: {filled_share}/{len(share_dict)} 只 → 第{DATA_START_ROW}行")

        # ── 市值（market_target_row, market_col）──
        filled_market = 0
        for code, market_val in market_dict.items():
            if code in code_to_col:
                market_col = code_to_col[code]
                ws.cell(market_target_row, market_col).value = float(market_val)
                filled_market += 1
        print(f"   💰 市值填入完成: {filled_market}/{len(market_dict)} 只 → 第{market_target_row}行")

        # ── 净值 / 月末净值 / 资金流动 ──
        filled_nav       = 0
        filled_nav_month = 0
        filled_cash_flow = 0

        for code, nav_val in nav_dict.items():
            if code not in code_to_col:
                continue
            market_col = code_to_col[code]
            share_col  = code_to_col[code] + 2

            ws.cell(ROW_NAV_DAILY, market_col).value = float(nav_val)
            filled_nav += 1

            if is_month_end:
                ws.cell(ROW_NAV_MONTHLY, market_col).value = float(nav_val)
                filled_nav_month += 1

            share_today     = ws.cell(DATA_START_ROW,     share_col).value
            share_yesterday = ws.cell(DATA_START_ROW + 1, share_col).value

            if share_today is not None and share_yesterday is not None:
                try:
                    share_growth = float(share_today) - float(share_yesterday)
                    cash_flow    = round(float(nav_val) * share_growth / 10000, 6)
                    ws.cell(ROW_CASH_FLOW, market_col).value = cash_flow
                    filled_cash_flow += 1
                    if filled_cash_flow <= 5 or filled_cash_flow % 10 == 0:
                        print(f"   💹 {code} 份额增长={share_growth:.2f}万份"
                              f"  资金流动={cash_flow:.4f}亿")
                except Exception as e:
                    print(f"   ⚠️ {code} 资金流动计算失败: {e}")
            elif share_today is not None and share_yesterday is None:
                print(f"   ⚠️ {code} 第19行（前日）为空，资金流动跳过")
            else:
                print(f"   ⚠️ {code} 第18行（当日）为空，资金流动跳过")

        print(f"   🔢 净值填入完成(第{ROW_NAV_DAILY}行): {filled_nav}/{len(nav_dict)} 只")
        if is_month_end:
            print(f"   🔢 月末净值填入完成(第{ROW_NAV_MONTHLY}行): {filled_nav_month}/{len(nav_dict)} 只")
        print(f"   💹 资金流动填入完成(第{ROW_CASH_FLOW}行): {filled_cash_flow}/{len(nav_dict)} 只")

        # ── 涨跌幅 第18行（change_col = market_col + 6）──
        filled_change = 0
        for code, chg_val in change_dict.items():
            if code in code_to_col:
                change_col = code_to_col[code] + 6
                cell = ws.cell(DATA_START_ROW, change_col)
                cell.value         = float(chg_val)
                cell.number_format = '0.00%'
                filled_change += 1
        print(f"   📊 涨跌幅填入完成: {filled_change}/{len(change_dict)} 只 → 第{DATA_START_ROW}行(+6列)")

        wb.save(EXCEL_FILE_PATH)
        total = len(SSE_LIST) + len(SZSE_LIST)
        print(f"\n✨ 全部完成！"
              f" 份额→第{DATA_START_ROW}行"
              f"  市值→第{market_target_row}行"
              f"  净值→第{ROW_NAV_DAILY}行"
              f"  资金流动→第{ROW_CASH_FLOW}行"
              f"  涨跌幅→第{DATA_START_ROW}行(+6列)"
              f"  ({len(share_dict)}/{total} 只)")

    except Exception as e:
        print(f"❌ 数据填入失败（请先关闭文件）: {e}")
        import traceback
        traceback.print_exc()


def run_integration():
    """主流程"""
    if not check_trading_day():
        return

    bj_today          = get_beijing_date()
    yesterday         = bj_today - timedelta(days=1)
    market_target_row = get_market_value_target_row(yesterday)
    is_month_end      = is_last_trading_day_of_month(yesterday)

    bj_now = datetime.utcnow() + timedelta(hours=8)
    print(f"{'='*55}")
    print(f"  ETF 份额采集  {bj_now.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*55}")

    print("\n【第一步】准备 Excel 新行...")
    code_to_col = prepare_excel_new_row()
    if code_to_col is None:
        print("❌ Excel 准备失败，程序终止")
        return

    print("\n【第二步】抓取数据...")
    szse_share, szse_market, szse_nav, szse_change = fetch_szse_data(SZSE_LIST)
    sse_share, sse_market, data_date, sse_nav, sse_change = fetch_sse_data(SSE_LIST)

    final_share = {**szse_share, **sse_share}
    all_market  = {**szse_market, **sse_market}
    all_nav     = {**szse_nav,    **sse_nav}
    all_change  = {**szse_change, **sse_change}

    print("\n【第三步】填入数据...")
    fill_excel_data(
        code_to_col, final_share, all_market, data_date,
        market_target_row, all_nav, is_month_end, all_change
    )

    elapsed = (datetime.now() - start_time).seconds
    total   = len(SSE_LIST) + len(SZSE_LIST)
    print(f"\n📊 采集汇总:"
          f"  份额 {len(final_share)}/{total} 只"
          f"  市值 {len(all_market)}/{total} 只"
          f"  净值 {len(all_nav)}/{total} 只"
          f"  涨跌幅 {len(all_change)}/{total} 只")
    print(f"⏱️  总耗时: {elapsed // 60} 分 {elapsed % 60} 秒")
    print(f"{'='*55}")


if __name__ == "__main__":
    start_time = datetime.now()
    run_integration()
